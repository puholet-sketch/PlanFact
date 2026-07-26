"""Сопоставление ФИО с командами из «Распределение по командам.xlsx»."""

from __future__ import annotations

import os
import re
from typing import Dict, List

from openpyxl import load_workbook

DEFAULT_MAPPING_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Распределение по командам.xlsx",
)
MAPPING_SHEET = "ВИРТУ СИСТЕМС"
SKIP_TEAMS = {"РГС ОФР"}
VFOS_TEAM = "РГС VFOS"

# Папка отчёта и заголовок (может отличаться от названия команды в файле распределения)
TEAM_OUTPUT_ALIASES = {
    "Все проекты": "Дизайнер",
}


def norm(value: str) -> str:
    text = (value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def safe_team_dir(team_name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", team_name).strip()


def output_team_name(team_name: str) -> str:
    return TEAM_OUTPUT_ALIASES.get(team_name, team_name)


def output_team_dir(team_name: str) -> str:
    return safe_team_dir(output_team_name(team_name))


def vfos_report_dir(source_root: str) -> str:
    return os.path.join(source_root, "teams", output_team_dir(VFOS_TEAM))


def load_team_rosters(path: str | None = None) -> Dict[str, List[str]]:
    """Команда -> список ФИО (как в файле распределения)."""
    path = path or DEFAULT_MAPPING_FILE
    if not os.path.exists(path):
        raise FileNotFoundError(f"Файл распределения не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if MAPPING_SHEET not in wb.sheetnames:
        raise ValueError(f"Лист «{MAPPING_SHEET}» не найден в {path}")

    rosters: Dict[str, List[str]] = {}
    ws = wb[MAPPING_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        fio = row[1] if len(row) > 1 else None
        team = row[4] if len(row) > 4 else None
        if not fio or not team:
            continue
        team = str(team).strip()
        fio = str(fio).strip()
        rosters.setdefault(team, []).append(fio)
    return rosters


def build_fio_index(rosters: Dict[str, List[str]]) -> Dict[str, str]:
    """Нормализованное ФИО -> команда."""
    index: Dict[str, str] = {}
    for team, fios in rosters.items():
        for fio in fios:
            index[norm(fio)] = team
    return index


def match_fio_to_team(fio: str, rosters: Dict[str, List[str]]) -> str | None:
    normalized = norm(fio)
    index = build_fio_index(rosters)
    if normalized in index:
        return index[normalized]
    parts = normalized.split()
    if len(parts) >= 2:
        for key, team in index.items():
            key_parts = key.split()
            if len(key_parts) >= 2 and parts[0] == key_parts[0] and parts[1][:1] == key_parts[1][:1]:
                return team
    return None
