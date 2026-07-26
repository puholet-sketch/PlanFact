"""Применить Jira-аудит к отчёту превышений."""

import json
import os
from collections import Counter, defaultdict

from openpyxl import Workbook, load_workbook

from jira_audit_rules import audit_task, format_reasons
from team_mapping import vfos_report_dir

SOURCE_ROOT = os.path.dirname(os.path.abspath(__file__))
VFOS_DIR = vfos_report_dir(SOURCE_ROOT)
PORTAL_URL = "https://portal.virtusystems.ru/browse"


def folder_paths(folder: str) -> dict[str, str]:
    return {
        "cache": os.path.join(folder, "jira_audit_cache.json"),
        "tasks": os.path.join(folder, "audit_tasks.json"),
        "report_v4": os.path.join(folder, "Превышения_по_ФИО_по_периодам_v4.xlsx"),
        "audit_xlsx": os.path.join(folder, "Аудит_нарушений_v1.xlsx"),
    }


def run_audit_for_folder(folder: str | None = None) -> tuple[list, dict]:
    folder = folder or SOURCE_ROOT
    fp = folder_paths(folder)
    tasks = json.load(open(fp["tasks"], encoding="utf-8"))
    cache = {}
    if os.path.exists(fp["cache"]):
        cache = json.load(open(fp["cache"], encoding="utf-8"))

    rows = []
    by_fio_violations: dict[str, Counter] = defaultdict(Counter)

    for key, meta in sorted(tasks.items()):
        jira = cache.get(key, {})
        audit = audit_task(
            key,
            meta["fio"],
            meta.get("work_type", ""),
            float(meta.get("plan") or 0),
            float(meta.get("fact") or 0),
            jira if jira else None,
        )
        for v in audit["violations"]:
            by_fio_violations[meta["fio"]][v] += 1

        rows.append(
            {
                "key": key,
                "fio": meta["fio"],
                "work_type": meta.get("work_type", ""),
                "plan": meta.get("plan"),
                "fact": meta.get("fact"),
                "delta": meta.get("delta"),
                "parent": audit.get("parent") or jira.get("parent"),
                "planning_keys": ", ".join(audit.get("planning_keys") or []),
                "jira_type": audit.get("jira_type") or jira.get("type"),
                "jira_checked": "да" if audit.get("checked") else "нет",
                "violations": "; ".join(audit["violations"]),
                "notes": "; ".join(audit["notes"]),
                "reasons": format_reasons(audit),
            }
        )

    return rows, by_fio_violations


def save_audit_xlsx(rows: list, by_fio_violations: dict, audit_xlsx: str | None = None) -> None:
    audit_xlsx = audit_xlsx or folder_paths(SOURCE_ROOT)["audit_xlsx"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Аудит задач"
    ws.append(
        [
            "Ключ",
            "ФИО",
            "Тип работы (выгрузка)",
            "План",
            "Факт",
            "Превышение",
            "Parent Story",
            "Planning keys",
            "Тип Jira",
            "Jira проверен",
            "Нарушения",
            "Заметки",
            "Итог причин",
        ]
    )
    for r in rows:
        ws.append(
            [
                r["key"],
                r["fio"],
                r["work_type"],
                r["plan"],
                r["fact"],
                r["delta"],
                r["parent"],
                r["planning_keys"],
                r["jira_type"],
                r["jira_checked"],
                r["violations"],
                r["notes"],
                r["reasons"],
            ]
        )
        cell = ws.cell(ws.max_row, 1)
        cell.hyperlink = f"{PORTAL_URL}/{r['key']}"
        cell.style = "Hyperlink"

    ws2 = wb.create_sheet("Свод по ФИО")
    ws2.append(["ФИО", "Нарушение", "Задач"])
    for fio in sorted(by_fio_violations.keys()):
        for violation, count in by_fio_violations[fio].most_common():
            ws2.append([fio, violation, count])

    wb.save(audit_xlsx)


def patch_report_v4(rows: list, folder: str | None = None) -> None:
    fp = folder_paths(folder or SOURCE_ROOT)
    report_v4 = fp["report_v4"]
    if not os.path.exists(report_v4):
        return
    reasons_by_key = {r["key"]: r["reasons"] for r in rows}
    wb = load_workbook(report_v4)
    if "Задачи с превышением" not in wb.sheetnames:
        wb.save(report_v4)
        return
    ws = wb["Задачи с превышением"]
    for row_idx in range(2, ws.max_row + 1):
        key = ws.cell(row_idx, 3).value
        if key in reasons_by_key:
            ws.cell(row_idx, 11).value = reasons_by_key[key]
    wb.save(report_v4)


def run_audit() -> tuple[list, dict]:
    return run_audit_for_folder(VFOS_DIR)


def main() -> None:
    rows, by_fio = run_audit()
    fp = folder_paths(VFOS_DIR)
    save_audit_xlsx(rows, by_fio, fp["audit_xlsx"])
    patch_report_v4(rows, VFOS_DIR)
    checked = sum(1 for r in rows if r["jira_checked"] == "да")
    print(f"AUDIT_XLSX: {fp['audit_xlsx']}")
    print(f"TASKS: {len(rows)} | Jira checked: {checked}")
    print(f"Updated: {fp['report_v4']}")


if __name__ == "__main__":
    main()
