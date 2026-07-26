import json
import os
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook, load_workbook

from jira_audit_rules import (
    CHECK_NEIGHBOR_PLANNING,
    DECOMP_RULES,
    JIRA_STRUCTURE_HINT,
    NEIGHBOR_PLANNING_HINT,
    REESTIMATE_EXAMPLE,
    REESTIMATE_STEPS,
    REESTIMATE_TITLE,
    RULES_TITLE,
    action_for_category,
    audit_task,
    categorize_violation,
    format_reasons,
)
from team_mapping import DEFAULT_MAPPING_FILE, VFOS_TEAM, load_team_rosters, vfos_report_dir

SOURCE_ROOT = os.path.dirname(os.path.abspath(__file__))
PORTAL_URL = "https://portal.virtusystems.ru/browse"

VFOS_RAW_TARGETS = [
    "Нижурин Сергей",
    "Агалетдинов Айнур",
    "Рудченко Павел",
    "Рыжикова Юлия",
    "Сивогривова Нина",
    "Кукарин Дмитрий",
]

KPI_COLUMN_HELP = [
    ("случаев превышения (ФИО × период)", "сколько раз у сотрудника команды в периоде Факт > План по категории «Производство»."),
    ("задач с превышением", "число строк в детализации — каждая строка = задача в конкретном периоде (см. «Строк задач»)."),
    ("суммарное превышение, ч", "сумма (Факт − План) по всем строкам задач с превышением за все периоды."),
    ("периодов в аналитике", "сколько файлов Plan-Fact (недель) включено в отчёт."),
]

FIO_TOTALS_COLUMN_HELP = [
    ("Периодов с превышением", "в скольких периодах у сотрудника суммарный Факт > План по «Производство»."),
    (
        "Строк задач",
        "сколько раз задача с превышением попала в детальный список; "
        "одна задача в 3 периодах = 3 строки.",
    ),
    (
        "Уник. задач",
        "число различных ключей Jira (RGS-…, SOGLVFOS-…); "
        "повторы в разных периодах считаются один раз.",
    ),
    ("Сумма +ч", "сумма часов превышения (Факт − План) по всем строкам задач сотрудника."),
]

VIOLATIONS_COLUMN_HELP = [
    ("Задач", "уникальных ключей Jira с данным типом нарушения."),
    ("+ч", "сумма превышения по этим задачам."),
    ("Story", "число родительских Story, затронутых нарушением."),
]

STORY_FEEDBACK_COLUMN_HELP = [
    ("Planning", "ключи задач «Планирование» / «Оценка» у Story (если есть в Jira)."),
    ("Задачи", "подзадачи с превышением, сгруппированные по Story."),
    ("+ч", "сумма превышения по Story."),
]

EMPLOYEE_TREND_COLUMN_HELP = [
    ("Коэффициент", "Факт / План по «Производство» за период (>1 — превышение)."),
    ("План → Факт", "суммарные часы по производству за период."),
    ("+ч", "разница Факт − План за период."),
]

FIO_TASKS_COLUMN_HELP = [
    ("Период", "дата среза из файла Plan-Fact."),
    ("+ч", "превышение по строке: Факт − План на подзадаче."),
    (
        "N задач в заголовке секции",
        "число строк в таблице ниже (не уникальных ключей; см. «Строк задач» в «Итого по ФИО»).",
    ),
]


@dataclass
class ReportConfig:
    source_folder: str
    output_folder: str
    team_name: str
    target_fios: list[str]
    use_jira_audit: bool = False


_CFG: ReportConfig | None = None


def default_vfos_config(rosters: dict | None = None) -> ReportConfig:
    targets = VFOS_RAW_TARGETS
    try:
        rosters = rosters or load_team_rosters(DEFAULT_MAPPING_FILE)
        targets = rosters.get(VFOS_TEAM, VFOS_RAW_TARGETS) or VFOS_RAW_TARGETS
    except FileNotFoundError:
        pass
    return ReportConfig(
        source_folder=SOURCE_ROOT,
        output_folder=vfos_report_dir(SOURCE_ROOT),
        team_name=VFOS_TEAM,
        target_fios=targets,
        use_jira_audit=True,
    )


def set_config(config: ReportConfig) -> None:
    global _CFG
    _CFG = config


def cfg() -> ReportConfig:
    global _CFG
    if _CFG is None:
        _CFG = default_vfos_config()
    return _CFG


def paths() -> dict[str, str]:
    out = cfg().output_folder
    return {
        "cache": os.path.join(out, "jira_audit_cache.json"),
        "audit_tasks": os.path.join(out, "audit_tasks.json"),
        "audit_xlsx": os.path.join(out, "Аудит_нарушений_v1.xlsx"),
        "story_feedback": os.path.join(out, "Обратная_связь_по_Story_v1.xlsx"),
        "report_xlsx": os.path.join(out, "Превышения_по_ФИО_по_периодам_v4.xlsx"),
        "report_html": os.path.join(out, "planfact_team_trend.html"),
    }


# Обратная совместимость
FOLDER = SOURCE_ROOT


def norm(value: str) -> str:
    text = (value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def is_target(name: str) -> bool:
    normalized = norm(name)
    for target in cfg().target_fios:
        target_norm = norm(target)
        if (
            normalized == target_norm
            or normalized.startswith(target_norm + " ")
            or target_norm.startswith(normalized + " ")
        ):
            return True
        if ("сивогривова" in normalized or "сиворгивова" in normalized) and (
            "сивогривова" in target_norm or "сиворгивова" in target_norm
        ):
            if "нина" in normalized and "нина" in target_norm:
                return True
        name_parts = normalized.split(" ")
        target_parts = target_norm.split(" ")
        if len(name_parts) >= 2 and len(target_parts) >= 2:
            if name_parts[0] == target_parts[0] and name_parts[1][:1] == target_parts[1][:1]:
                return True
    return False


def get_period_from_filename(path: str) -> str:
    match = re.search(r"(\d{4}\.\d{2}\.\d{2})", os.path.basename(path))
    return match.group(1) if match else "unknown"


def to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def is_production_row(row_vals) -> bool:
    for idx in (1, 4):
        if len(row_vals) >= idx + 1:
            candidate = row_vals[idx]
            if isinstance(candidate, str) and "производ" in norm(candidate):
                return True
    return False


def list_source_files() -> list[str]:
    files = []
    folder = cfg().source_folder
    for name in os.listdir(folder):
        lower = name.lower()
        if not lower.endswith(".xlsx"):
            continue
        if name.startswith("~$") or name.startswith("Превышения_"):
            continue
        if name.startswith("Аудит_") or name.startswith("Обратная_"):
            continue
        if lower in ("planfact_team_trend.html", "audit_tasks.json", "jira_audit_cache.json"):
            continue
        if not re.search(r"\d{4}\.\d{2}\.\d{2}", name):
            continue
        if "модель" not in lower and "план-факт" not in lower and "plan-fact" not in lower:
            continue
        files.append(os.path.join(folder, name))
    return sorted(files)


def classify_exceed_reasons(
    fio: str, task_key: str, work_type: str, project: str, plan_h: float, fact_h: float
) -> list[str]:
    """Эвристика причин. Не утверждаем «нет задачи Планирование» без проверки Jira."""
    reasons = []
    wt = norm(work_type)
    proj = norm(str(project or ""))
    key = norm(str(task_key or ""))
    delta = fact_h - plan_h

    if plan_h == 0:
        reasons.append(
            "На подзадаче план = 0 ч (превышение относительно нуля); "
            f"{NEIGHBOR_PLANNING_HINT}"
        )
        reasons.append("Аналитикам — не забывать проставлять оценки")

    elif plan_h > 0 and fact_h > plan_h:
        if any(token in wt for token in ("анализ", "сопровожд", "документ", "исследован")):
            reasons.append(
                f"Неточная оценка на подзадаче аналитики/сопровождения "
                f"(план {plan_h:g} ч → факт {fact_h:g} ч, +{delta:g} ч); "
                f"{CHECK_NEIGHBOR_PLANNING}"
            )
        elif fact_h >= plan_h * 1.35:
            reasons.append(
                f"Существенный перерасход при наличии оценки (+{delta:g} ч) — "
                f"вероятны новые требования или расширение объёма работ"
            )
        else:
            reasons.append(
                f"Промах в оценке: план {plan_h:g} ч, факт {fact_h:g} ч (+{delta:g} ч)"
            )

    if (
        plan_h == 16
        or "external" in key
        or "external" in proj
        or "дефект" in wt
        or "bug" in key
    ):
        reasons.append("External / дефект с прода (стандарт 16 ч или внешний инцидент)")

    size = max(plan_h, fact_h)
    if size > 24:
        reasons.append("Задача >24 ч — нарушена гранулярность, нужно дробление BA/DEV/QA")
    elif size > 20:
        reasons.append("Крупная задача (>20 ч) — риск превышения без дробления")

    if not reasons:
        reasons.append("Факт превысил план — разбор с аналитиком / ПМ / тимлидом")

    return reasons


def load_jira_cache() -> dict:
    if not cfg().use_jira_audit:
        return {}
    cache_file = paths()["cache"]
    if not os.path.exists(cache_file):
        return {}
    return json.load(open(cache_file, encoding="utf-8"))


def get_task_reasons(
    fio: str,
    task_key: str,
    work_type: str,
    project: str,
    plan_h: float,
    fact_h: float,
    jira_cache: dict,
) -> list[str]:
    jira = jira_cache.get(str(task_key or "").strip(), {})
    if jira:
        audit = audit_task(fio, task_key, work_type, plan_h, fact_h, jira)
        text = format_reasons(audit)
        return [part.strip() for part in text.split("; ") if part.strip()]
    return classify_exceed_reasons(fio, task_key, work_type, project, plan_h, fact_h)


def export_audit_tasks(task_rows: list) -> None:
    """Уникальные ключи с максимальным превышением для Jira-аудита."""
    tasks: dict[str, dict] = {}
    for row in task_rows:
        key = str(row[2] or "").strip()
        if not key:
            continue
        delta = float(row[7] or 0)
        entry = {
            "fio": row[1],
            "work_type": row[3],
            "plan": row[5],
            "fact": row[6],
            "delta": delta,
        }
        if key not in tasks or delta > tasks[key]["delta"]:
            tasks[key] = entry
    json.dump(tasks, open(paths()["audit_tasks"], "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def save_audit_sheet(wb_out, jira_cache: dict, task_rows: list) -> None:
    if not jira_cache:
        return
    ws = wb_out.create_sheet("Аудит нарушений")
    ws.append(
        [
            "Ключ",
            "ФИО",
            "План",
            "Факт",
            "Превышение",
            "Parent",
            "Planning",
            "Jira тип",
            "Нарушения",
            "Заметки",
        ]
    )
    seen = set()
    for row in task_rows:
        key = str(row[2] or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        jira = jira_cache.get(key, {})
        audit = audit_task(row[1], key, row[3], float(row[5]), float(row[6]), jira or None)
        ws.append(
            [
                key,
                row[1],
                row[5],
                row[6],
                row[7],
                audit.get("parent") or jira.get("parent"),
                ", ".join(audit.get("planning_keys") or []),
                audit.get("jira_type") or jira.get("type"),
                "; ".join(audit["violations"]),
                "; ".join(audit["notes"]),
            ]
        )
        cell = ws.cell(ws.max_row, 1)
        cell.hyperlink = f"{PORTAL_URL}/{key}"
        cell.style = "Hyperlink"


def collect_data(jira_cache: dict | None = None) -> tuple[list, list, dict]:
    jira_cache = jira_cache if jira_cache is not None else load_jira_cache()
    summary_rows = []
    task_rows = []
    trend = {
        "periods": [],
        "people": {},
        "team_by_period": {},
    }

    for path in list_source_files():
        period = get_period_from_filename(path)
        if period not in trend["periods"]:
            trend["periods"].append(period)

        wb = load_workbook(path, data_only=True, read_only=True)
        if "Исходные данные" not in wb.sheetnames:
            continue

        ws_src = wb["Исходные данные"]
        person_totals = {}
        person_display = {}
        source_rows = []

        for row_vals in ws_src.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):
            if not is_production_row(row_vals):
                continue

            fio = row_vals[0] if len(row_vals) >= 1 else None
            if not isinstance(fio, str):
                continue
            fio = fio.strip()
            if not fio or not is_target(fio):
                continue

            fio_key = norm(fio)
            plan_h = to_float(row_vals[8] if len(row_vals) >= 9 else None)
            fact_h = to_float(row_vals[9] if len(row_vals) >= 10 else None)

            if fio_key not in person_totals:
                person_totals[fio_key] = {"plan": 0.0, "fact": 0.0}
                person_display[fio_key] = fio
            person_totals[fio_key]["plan"] += plan_h
            person_totals[fio_key]["fact"] += fact_h
            source_rows.append((fio_key, row_vals, plan_h, fact_h))

        exceeded = {}
        for fio_key, totals in person_totals.items():
            plan = totals["plan"]
            fact = totals["fact"]
            excess = fact - plan
            ratio = fact / plan if plan else 0.0
            display_name = person_display[fio_key]

            if fio_key not in trend["people"]:
                trend["people"][fio_key] = {"name": display_name, "periods": {}}
            trend["people"][fio_key]["periods"][period] = {
                "plan": round(plan, 2),
                "fact": round(fact, 2),
                "excess": round(excess, 2),
                "ratio": round(ratio, 4),
                "has_exceed": fact > plan,
                "tasks": [],
            }

            if period not in trend["team_by_period"]:
                trend["team_by_period"][period] = {
                    "tasks_exceed": 0,
                    "hours_exceed": 0.0,
                    "people_exceed": 0,
                }

            if fact > plan:
                exceeded[fio_key] = display_name
                trend["team_by_period"][period]["people_exceed"] += 1
                summary_rows.append(
                    [
                        period,
                        display_name,
                        round(plan, 2),
                        round(fact, 2),
                        round(ratio, 4),
                        round(excess, 2),
                        os.path.basename(path),
                    ]
                )

        for fio_key, row_vals, plan_h, fact_h in source_rows:
            if fio_key not in exceeded:
                continue
            delta = fact_h - plan_h
            if delta <= 0:
                continue

            date_val = row_vals[10] if len(row_vals) >= 11 else None
            if isinstance(date_val, datetime):
                date_val = date_val.date().isoformat()

            task_key = (row_vals[3] if len(row_vals) >= 4 else "") or ""
            work_type = (row_vals[2] if len(row_vals) >= 3 else "") or ""
            category = (row_vals[1] if len(row_vals) >= 2 else "") or ""
            project = (row_vals[4] if len(row_vals) >= 5 else "") or ""
            display_name = exceeded[fio_key]

            reasons = get_task_reasons(
                display_name, task_key, work_type, project, plan_h, fact_h, jira_cache
            )

            task_row = [
                period,
                display_name,
                task_key,
                work_type,
                category,
                round(plan_h, 2),
                round(fact_h, 2),
                round(delta, 2),
                date_val,
                os.path.basename(path),
                "; ".join(reasons),
            ]
            task_rows.append(task_row)
            trend["team_by_period"][period]["tasks_exceed"] += 1
            trend["team_by_period"][period]["hours_exceed"] += delta
            trend["people"][fio_key]["periods"][period]["tasks"].append(
                {
                    "key": task_key,
                    "type": work_type,
                    "project": project,
                    "plan": round(plan_h, 2),
                    "fact": round(fact_h, 2),
                    "excess": round(delta, 2),
                    "date": date_val,
                    "reasons": reasons,
                }
            )

    trend["periods"] = sorted(trend["periods"])
    for period in trend["periods"]:
        vals = trend["team_by_period"].setdefault(
            period, {"tasks_exceed": 0, "hours_exceed": 0.0, "people_exceed": 0}
        )
        vals["hours_exceed"] = round(vals["hours_exceed"], 2)

    summary_rows.sort(key=lambda row: (row[0], row[1]))
    task_rows.sort(key=lambda row: (row[0], row[1], -row[7], str(row[2])))
    return summary_rows, task_rows, trend


def aggregate_fio_reasons(task_rows: list) -> dict[str, list[tuple[str, int, float]]]:
    grouped: dict[str, Counter] = {}
    hours: dict[str, dict[str, float]] = {}

    for row in task_rows:
        fio = row[1]
        for reason in str(row[10]).split("; "):
            reason = reason.strip()
            if not reason:
                continue
            grouped.setdefault(fio, Counter())[reason] += 1
            hours.setdefault(fio, {}).setdefault(reason, 0.0)
            hours[fio][reason] += row[7]

    result = {}
    for fio, counter in grouped.items():
        top = []
        for reason, count in counter.most_common(5):
            top.append((reason, count, round(hours[fio][reason], 1)))
        result[fio] = top
    return result


def build_fio_totals(task_rows: list, summary_rows: list) -> tuple[list[dict], dict]:
    """Итого по ФИО: периоды с превышением, строки задач, уникальные ключи, часы."""
    cases_by_fio: Counter = Counter(row[1] for row in summary_rows)
    totals: dict[str, dict] = {}

    for row in task_rows:
        fio = row[1]
        entry = totals.setdefault(
            fio,
            {"fio": fio, "task_rows": 0, "unique_keys": set(), "hours": 0.0},
        )
        entry["task_rows"] += 1
        entry["hours"] += float(row[7] or 0)
        key = str(row[2] or "").strip()
        if key:
            entry["unique_keys"].add(key)

    rows = []
    grand = {"periods": 0, "task_rows": 0, "unique_keys": 0, "hours": 0.0}
    for fio in sorted(totals.keys()):
        t = totals[fio]
        periods = cases_by_fio.get(fio, 0)
        unique = len(t["unique_keys"])
        hours = round(t["hours"], 1)
        rows.append(
            {
                "fio": fio,
                "periods": periods,
                "task_rows": t["task_rows"],
                "unique_keys": unique,
                "hours": hours,
            }
        )
        grand["periods"] += periods
        grand["task_rows"] += t["task_rows"]
        grand["unique_keys"] += unique
        grand["hours"] += hours
    grand["hours"] = round(grand["hours"], 1)
    return rows, grand


def build_audit_aggregates(
    task_rows: list, jira_cache: dict
) -> tuple[list, list, list]:
    """Свод по типам нарушений и группировка по Story для обратной связи."""
    type_stats: dict[str, dict] = {}
    seen_keys: set[str] = set()
    story_map: dict[str, dict] = {}

    for row in task_rows:
        key = str(row[2] or "").strip()
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)

        jira = jira_cache.get(key, {})
        audit = audit_task(row[1], key, row[3], float(row[5]), float(row[6]), jira or None)
        delta = float(row[7] or 0)
        parent = audit.get("parent") or jira.get("parent") or key
        story_key = str(parent or key)

        violation_categories: set[str] = set()
        for v in audit["violations"]:
            violation_categories.add(categorize_violation(v))
        for part in str(row[10] if len(row) > 10 else "").split("; "):
            part = part.strip()
            if part and not part.startswith("[Jira"):
                violation_categories.add(categorize_violation(part))

        if story_key not in story_map:
            story_map[story_key] = {
                "story": story_key,
                "planning": set(),
                "categories": set(),
                "tasks": set(),
                "executors": set(),
                "hours": 0.0,
                "violations_raw": set(),
            }
        story = story_map[story_key]
        story["tasks"].add(key)
        story["executors"].add(row[1])
        story["hours"] += delta
        for pk in audit.get("planning_keys") or []:
            if pk:
                story["planning"].add(pk)
        for cat in violation_categories:
            story["categories"].add(cat)
            stats = type_stats.setdefault(
                cat, {"category": cat, "tasks": set(), "hours": 0.0, "stories": set()}
            )
            stats["tasks"].add(key)
            stats["hours"] += delta
            stats["stories"].add(story_key)
        for v in audit["violations"]:
            story["violations_raw"].add(v)

    type_rows = []
    for cat, stats in sorted(type_stats.items(), key=lambda x: (-len(x[1]["tasks"]), -x[1]["hours"])):
        type_rows.append(
            {
                "category": cat,
                "tasks_count": len(stats["tasks"]),
                "hours": round(stats["hours"], 1),
                "stories_count": len(stats["stories"]),
                "action": action_for_category(cat),
            }
        )

    story_rows = []
    for story_key in sorted(story_map.keys(), key=lambda k: -story_map[k]["hours"]):
        s = story_map[story_key]
        cats = sorted(s["categories"])
        story_rows.append(
            {
                "story": story_key,
                "planning": ", ".join(sorted(s["planning"])),
                "categories": cats,
                "categories_text": "; ".join(cats),
                "tasks": sorted(s["tasks"]),
                "tasks_text": ", ".join(sorted(s["tasks"])),
                "executors": ", ".join(sorted(s["executors"])),
                "hours": round(s["hours"], 1),
                "actions": "; ".join(dict.fromkeys(action_for_category(c) for c in cats)),
            }
        )

    return type_rows, story_rows, list(seen_keys)


def save_story_feedback_xlsx(type_rows: list, story_rows: list) -> None:
    wb = Workbook()

    ws_types = wb.active
    ws_types.title = "Нарушения по типам"
    ws_types.append(["Тип нарушения", "Задач", "Часы превышения", "Story", "Рекомендуемое действие"])
    for row in type_rows:
        ws_types.append(
            [row["category"], row["tasks_count"], row["hours"], row["stories_count"], row["action"]]
        )

    ws_stories = wb.create_sheet("Обратная связь по Story")
    ws_stories.append(
        [
            "Story",
            "Planning",
            "Типы нарушений",
            "Задачи с превышением",
            "Исполнители",
            "Сумма превышения, ч",
            "Рекомендуемые действия",
        ]
    )
    for row in story_rows:
        ws_stories.append(
            [
                row["story"],
                row["planning"],
                row["categories_text"],
                row["tasks_text"],
                row["executors"],
                row["hours"],
                row["actions"],
            ]
        )
        cell = ws_stories.cell(ws_stories.max_row, 1)
        cell.hyperlink = f"{PORTAL_URL}/{row['story']}"
        cell.style = "Hyperlink"

    wb.save(paths()["story_feedback"])


def save_xlsx(summary_rows: list, task_rows: list, jira_cache: dict | None = None) -> None:
    jira_cache = jira_cache if jira_cache is not None else load_jira_cache()
    wb_out = Workbook()
    ws_summary = wb_out.active
    ws_summary.title = "Свод превышений"
    ws_summary.append(
        ["Период", "ФИО", "План", "Факт", "Превышение (коэф.)", "Превышение (часы)", "Файл"]
    )
    for row in summary_rows:
        ws_summary.append(row)

    ws_tasks = wb_out.create_sheet("Задачи с превышением")
    ws_tasks.append(
        [
            "Период",
            "ФИО",
            "Ключ задачи",
            "Тип работы",
            "Категория",
            "План (ч)",
            "Факт (ч)",
            "Превышение (ч)",
            "Дата",
            "Файл",
            "Вероятные причины",
        ]
    )
    for row in task_rows:
        ws_tasks.append(row)

    for row_idx in range(2, ws_tasks.max_row + 1):
        task_key = ws_tasks.cell(row_idx, 3).value
        if isinstance(task_key, str) and task_key.strip():
            key = task_key.strip()
            ws_tasks.cell(row_idx, 3).hyperlink = f"{PORTAL_URL}/{key}"
            ws_tasks.cell(row_idx, 3).style = "Hyperlink"

    fio_reasons = aggregate_fio_reasons(task_rows)
    ws_reasons = wb_out.create_sheet("Причины по ФИО")
    ws_reasons.append(["ФИО", "Причина", "Задач", "Часы превышения"])
    for fio in sorted(fio_reasons.keys()):
        for reason, count, hrs in fio_reasons[fio]:
            ws_reasons.append([fio, reason, count, hrs])

    ws_top = wb_out.create_sheet("Топ причин")
    ws_top.append(["Период", "ФИО", "Ключ задачи", "Превышение (ч)", "План (ч)", "Факт (ч)"])
    top_items = {}
    for row in task_rows:
        key = (row[0], row[1])
        if key not in top_items or row[7] > top_items[key][7]:
            top_items[key] = row
    for key in sorted(top_items.keys()):
        row = top_items[key]
        ws_top.append([row[0], row[1], row[2], row[7], row[5], row[6]])
        row_idx = ws_top.max_row
        task_key = ws_top.cell(row_idx, 3).value
        if isinstance(task_key, str) and task_key.strip():
            ws_top.cell(row_idx, 3).hyperlink = f"{PORTAL_URL}/{task_key.strip()}"
            ws_top.cell(row_idx, 3).style = "Hyperlink"

    save_audit_sheet(wb_out, jira_cache, task_rows)

    fio_total_rows, _ = build_fio_totals(task_rows, summary_rows)
    ws_fio = wb_out.create_sheet("Итого по ФИО")
    ws_fio.append(
        [
            "ФИО",
            "Периодов с превышением",
            "Строк задач с превышением",
            "Уникальных задач",
            "Сумма превышения, ч",
        ]
    )
    for row in fio_total_rows:
        ws_fio.append(
            [row["fio"], row["periods"], row["task_rows"], row["unique_keys"], row["hours"]]
        )

    if task_rows:
        type_rows, _, _ = build_audit_aggregates(task_rows, jira_cache or {})
        ws_v = wb_out.create_sheet("Нарушения по типам")
        ws_v.append(
            ["Тип нарушения", "Задач", "Часы превышения", "Story", "Рекомендуемое действие"]
        )
        for row in type_rows:
            ws_v.append(
                [
                    row["category"],
                    row["tasks_count"],
                    row["hours"],
                    row["stories_count"],
                    row["action"],
                ]
            )

    save_reference_sheet(wb_out)
    wb_out.save(paths()["report_xlsx"])


def save_reference_sheet(wb_out) -> None:
    ws = wb_out.create_sheet("Справка", 0)
    ws.append([RULES_TITLE])
    for rule in DECOMP_RULES:
        ws.append(["", rule])
    ws.append([])
    ws.append([REESTIMATE_TITLE])
    for idx, step in enumerate(REESTIMATE_STEPS, 1):
        ws.append(["", f"{idx}. {step}"])
    ws.append(["", REESTIMATE_EXAMPLE])
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100


def esc(text) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def column_legend_html(items: list[tuple[str, str]], intro: str = "") -> str:
    lis = "".join(
        f"<li><strong>{esc(title)}</strong> — {esc(text)}</li>" for title, text in items
    )
    intro_html = f"<p class='muted'>{esc(intro)}</p>" if intro else ""
    return f"<div class='column-legend'>{intro_html}<ul>{lis}</ul></div>"


def slugify_team(name: str) -> str:
    text = (name or "").strip().lower().replace("ё", "е")
    text = re.sub(r"[^a-z0-9а-я]+", "-", text, flags=re.IGNORECASE)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "team"


def build_team_payload(
    trend: dict,
    summary_rows: list,
    task_rows: list,
    jira_cache: dict | None = None,
) -> dict:
    """Сериализуемое представление команды для единого HTML."""
    jira_cache = jira_cache if jira_cache is not None else load_jira_cache()
    periods = trend["periods"]
    team = trend["team_by_period"]
    fio_reasons = aggregate_fio_reasons(task_rows)

    type_rows: list = []
    if task_rows:
        type_rows, _, _ = build_audit_aggregates(task_rows, jira_cache or {})

    chart_labels = periods
    chart_tasks = [team.get(p, {}).get("tasks_exceed", 0) for p in periods]
    chart_hours = [round(team.get(p, {}).get("hours_exceed", 0.0), 1) for p in periods]
    chart_people = [team.get(p, {}).get("people_exceed", 0) for p in periods]

    total_excess = round(sum(chart_hours), 1)
    total_tasks = len(task_rows)
    total_cases = len(summary_rows)
    unique_keys = len({str(r[2]).strip() for r in task_rows if r[2]})

    last_period = periods[-1] if periods else None
    prev_period = periods[-2] if len(periods) >= 2 else None
    last_hours = float(team.get(last_period, {}).get("hours_exceed", 0) or 0) if last_period else 0.0
    prev_hours = float(team.get(prev_period, {}).get("hours_exceed", 0) or 0) if prev_period else 0.0
    hours_delta = round(last_hours - prev_hours, 1) if prev_period else None

    people_trend = []
    for target in cfg().target_fios:
        person = None
        for pdata in trend["people"].values():
            if norm(pdata["name"]) == norm(target) or norm(pdata["name"]).startswith(
                norm(target)
            ):
                person = pdata
                break
        if not person:
            for pdata in trend["people"].values():
                if norm(target).split()[0] in norm(pdata["name"]):
                    person = pdata
                    break
        if not person:
            continue

        cells = []
        for period in periods:
            pdata = person["periods"].get(period)
            if not pdata:
                cells.append(None)
                continue
            cells.append(
                {
                    "ratio": pdata["ratio"],
                    "plan": pdata["plan"],
                    "fact": pdata["fact"],
                    "excess": pdata["excess"],
                    "has_exceed": pdata["has_exceed"],
                }
            )
        people_trend.append({"name": person["name"], "cells": cells})

    fio_total_rows, fio_grand = build_fio_totals(task_rows, summary_rows)
    top_person = fio_total_rows[0] if fio_total_rows else None
    concentration = 0.0
    if fio_grand["hours"] and len(fio_total_rows) >= 2:
        concentration = round(
            100.0
            * (float(fio_total_rows[0]["hours"]) + float(fio_total_rows[1]["hours"]))
            / float(fio_grand["hours"]),
            1,
        )
    elif fio_grand["hours"] and fio_total_rows:
        concentration = 100.0

    tasks_by_fio = []
    grouped: dict[str, list] = {}
    for row in task_rows:
        grouped.setdefault(row[1], []).append(row)
    for fio in sorted(grouped.keys()):
        reason_items = [
            {"reason": reason, "count": count, "hours": hrs}
            for reason, count, hrs in fio_reasons.get(fio, [])
        ]
        rows = []
        for row in grouped[fio]:
            key = str(row[2] or "").strip()
            rows.append(
                {
                    "period": row[0],
                    "key": key,
                    "url": f"{PORTAL_URL}/{key}" if key else "#",
                    "work_type": row[3],
                    "plan": row[5],
                    "fact": row[6],
                    "excess": row[7],
                    "reasons": row[10],
                }
            )
        tasks_by_fio.append(
            {
                "fio": fio,
                "count": len(grouped[fio]),
                "hours": round(sum(r[7] for r in grouped[fio]), 1),
                "reasons": reason_items,
                "rows": rows,
            }
        )

    violations = [
        {
            "category": row["category"],
            "tasks_count": row["tasks_count"],
            "hours": row["hours"],
            "stories_count": row["stories_count"],
            "action": row["action"],
        }
        for row in type_rows
    ]

    return {
        "slug": slugify_team(cfg().team_name),
        "name": cfg().team_name,
        "periods": periods,
        "kpi": {
            "people": len(cfg().target_fios),
            "cases": total_cases,
            "tasks": total_tasks,
            "hours": total_excess,
            "unique_keys": unique_keys,
            "top_fio": top_person["fio"] if top_person else "",
        },
        "chart": {
            "labels": chart_labels,
            "tasks": chart_tasks,
            "hours": chart_hours,
            "people": chart_people,
        },
        "fio_totals": [
            {
                "fio": r["fio"],
                "periods": r["periods"],
                "task_rows": r["task_rows"],
                "unique_keys": r["unique_keys"],
                "hours": r["hours"],
            }
            for r in fio_total_rows
        ],
        "fio_grand": {
            "periods": fio_grand["periods"],
            "task_rows": fio_grand["task_rows"],
            "unique_keys": fio_grand["unique_keys"],
            "hours": fio_grand["hours"],
        },
        "people_trend": people_trend,
        "violations": violations,
        "tasks_by_fio": tasks_by_fio,
        "insights": {
            "last_period": last_period,
            "prev_period": prev_period,
            "last_hours": round(last_hours, 1),
            "hours_delta": hours_delta,
            "top_fio": top_person["fio"] if top_person else "",
            "top_hours": top_person["hours"] if top_person else 0,
            "concentration": concentration,
        },
    }


def save_html(
    trend: dict,
    summary_rows: list,
    task_rows: list,
    jira_cache: dict | None = None,
) -> None:
    payload = build_team_payload(trend, summary_rows, task_rows, jira_cache)
    from report_theme import render_unified_site

    html = render_unified_site(
        {
            "periods": payload["periods"],
            "generated": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "source_count": len(payload["periods"]),
            "rules": DECOMP_RULES,
            "reestimate_steps": REESTIMATE_STEPS,
            "rules_title": RULES_TITLE,
            "reestimate_title": REESTIMATE_TITLE,
            "reestimate_example": REESTIMATE_EXAMPLE,
        },
        [payload],
    )
    with open(paths()["report_html"], "w", encoding="utf-8") as handle:
        handle.write(html)


def build_report(config: ReportConfig | None = None) -> dict | None:
    if config is not None:
        set_config(config)
    report_paths = paths()
    output_folder = cfg().output_folder

    if output_folder.upper().startswith("D:") is False and "PlanFact" not in output_folder:
        print(f"WARNING: нестандартная папка вывода: {output_folder}")

    files = list_source_files()
    if not files:
        print(f"Нет исходных файлов в {cfg().source_folder}")
        return None

    jira_cache = load_jira_cache()
    summary_rows, task_rows, trend = collect_data(jira_cache)
    export_audit_tasks(task_rows)
    save_xlsx(summary_rows, task_rows, jira_cache)
    payload = build_team_payload(trend, summary_rows, task_rows, jira_cache)
    from report_theme import render_unified_site

    html = render_unified_site(
        {
            "periods": payload["periods"],
            "generated": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "source_count": len(payload["periods"]),
            "rules": DECOMP_RULES,
            "reestimate_steps": REESTIMATE_STEPS,
            "rules_title": RULES_TITLE,
            "reestimate_title": REESTIMATE_TITLE,
            "reestimate_example": REESTIMATE_EXAMPLE,
        },
        [payload],
    )
    with open(report_paths["report_html"], "w", encoding="utf-8") as handle:
        handle.write(html)

    type_rows: list = []
    story_rows: list = []
    if task_rows:
        type_rows, story_rows, _ = build_audit_aggregates(task_rows, jira_cache or {})
        save_story_feedback_xlsx(type_rows, story_rows)
        print(f"STORY_FEEDBACK_XLSX: {report_paths['story_feedback']}")

    if jira_cache and cfg().use_jira_audit:
        from apply_jira_audit import run_audit_for_folder, save_audit_xlsx

        audit_rows, by_fio = run_audit_for_folder(output_folder)
        save_audit_xlsx(audit_rows, by_fio, report_paths["audit_xlsx"])
        print(
            f"AUDIT_XLSX: {report_paths['audit_xlsx']} "
            f"({len(audit_rows)} задач, Jira: {len(jira_cache)})"
        )
    elif cfg().use_jira_audit:
        print(f"AUDIT: cache missing ({report_paths['cache']}) — только эвристика")
    else:
        print("AUDIT: Jira-аудит отключён для этой команды")

    print(f"TEAM: {cfg().team_name}")
    print(f"OUTPUT: {output_folder}")
    print(f"SOURCE_FILES: {len(files)}")
    for path in files:
        print(f"  - {os.path.basename(path)}")
    print(f"REPORT_XLSX: {report_paths['report_xlsx']}")
    print(f"REPORT_HTML: {report_paths['report_html']}")
    print(f"SUMMARY_ROWS: {len(summary_rows)}")
    print(f"TASK_ROWS: {len(task_rows)}")
    return payload


def main() -> None:
    build_report(default_vfos_config())


if __name__ == "__main__":
    main()
